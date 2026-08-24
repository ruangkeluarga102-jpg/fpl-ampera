"""
FPL Exporter Module
Handles formatted exports to Excel (.xlsx), CSV, and JSON.
"""

import pandas as pd
from typing import Dict, Any, Optional
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class FPLExporter:
    def __init__(self, league_data: Dict[str, Any]):
        self.data = league_data
        self.league_name = league_data.get("league_info", {}).get("name", "FPL_Mini_League")
        self.gameweek = league_data.get("gameweek", 1)

    def _sanitize_filename(self, name: str) -> str:
        return "".join(c for c in name if c.isalnum() or c in (" ", "_", "-")).rstrip().replace(" ", "_")

    def export_to_excel(self, output_path: Optional[str] = None) -> str:
        """
        Export complete mini-league analysis to a beautifully styled multi-sheet Excel file.
        """
        if not output_path:
            clean_name = self._sanitize_filename(self.league_name)
            output_path = f"FPL_Report_{clean_name}_GW{self.gameweek}.xlsx"

        # Prepare Clean DataFrames for Export
        standings_export = self.data.get("standings_df", pd.DataFrame()).copy()
        # Drop internal fields if present
        cols_to_drop = [c for c in ["entry_id", "squad"] if c in standings_export.columns]
        if cols_to_drop:
            standings_export = standings_export.drop(columns=cols_to_drop)

        ownership_export = self.data.get("ownership_df", pd.DataFrame()).copy()
        captaincy_export = self.data.get("captaincy_df", pd.DataFrame()).copy()
        chips_export = self.data.get("chips_df", pd.DataFrame()).copy()
        history_export = self.data.get("history_df", pd.DataFrame()).copy()
        transfers_export = self.data.get("transfers_df", pd.DataFrame()).copy()

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            standings_export.to_excel(writer, sheet_name="Standings", index=False)
            ownership_export.to_excel(writer, sheet_name="Ownership_EO", index=False)
            captaincy_export.to_excel(writer, sheet_name="Captain_Picks", index=False)
            chips_export.to_excel(writer, sheet_name="Chip_Tracker", index=False)
            if not transfers_export.empty:
                cols_to_drop = [c for c in ["entry_id"] if c in transfers_export.columns]
                if cols_to_drop:
                    transfers_export = transfers_export.drop(columns=cols_to_drop)
                transfers_export.to_excel(writer, sheet_name="Transfers_GW", index=False)
            if not history_export.empty:
                history_export.to_excel(writer, sheet_name="GW_History", index=False)

        # Apply Openpyxl Styling
        wb = openpyxl.load_workbook(output_path)
        
        # Color Palette
        header_fill = PatternFill(start_color="37003C", end_color="37003C", fill_type="solid") # Premier League Purple
        header_font = Font(name="Calibri", size=11, bold=True, color="00FF87") # Premier League Cyan/Green
        regular_font = Font(name="Calibri", size=10)
        border_thin = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            ws.views.sheetView[0].showGridLines = True
            
            # Format header row
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Format data rows & auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row > 1:
                        cell.font = regular_font
                        cell.border = border_thin
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    val_str = str(cell.value or "")
                    max_len = max(max_len, len(val_str))
                
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
            # Freeze Header Row
            ws.freeze_panes = "A2"

        wb.save(output_path)
        return os.path.abspath(output_path)

    def export_to_csv(self, base_filename: Optional[str] = None) -> str:
        """Export standings and ownership to CSV."""
        if not base_filename:
            clean_name = self._sanitize_filename(self.league_name)
            base_filename = f"FPL_Standings_{clean_name}_GW{self.gameweek}.csv"
        
        standings_df = self.data.get("standings_df", pd.DataFrame()).copy()
        cols_to_drop = [c for c in ["entry_id", "squad"] if c in standings_df.columns]
        if cols_to_drop:
            standings_df = standings_df.drop(columns=cols_to_drop)
            
        standings_df.to_csv(base_filename, index=False)
        return os.path.abspath(base_filename)
